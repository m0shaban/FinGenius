from flask import render_template, make_response, request, current_app, send_file, jsonify
from flask_login import login_required, current_user
import io
import csv
import tempfile
import os
from datetime import datetime
from app.export import bp
from app.core.models import FinancialFile, Analysis
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

@bp.route('/pdf/analysis/<int:file_id>')
@login_required
def export_analysis_pdf(file_id):
    try:
        file = FinancialFile.query.get_or_404(file_id)
        if file.user_id != current_user.id:
            flash('Access denied')
            return redirect(url_for('core.dashboard'))
        
        # Get latest analysis
        analysis = Analysis.query.filter_by(file_id=file_id).order_by(Analysis.created_date.desc()).first()
        
        if not analysis:
            flash('No analysis found for this file')
            return redirect(url_for('core.view_file', file_id=file_id))
        
        # Generate PDF
        pdf = ExportService.generate_pdf_report(
            'export/analysis_report.html',
            file=file,
            analysis=analysis,
            date=datetime.now()
        )
        
        return send_file(
            pdf,
            download_name=f'analysis_{file.filename}_{datetime.now():%Y%m%d}.pdf',
            as_attachment=True,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}')
        return redirect(url_for('core.view_file', file_id=file_id))

@bp.route('/excel/analysis/<int:file_id>')
@login_required
def export_analysis_excel(file_id):
    try:
        file = FinancialFile.query.get_or_404(file_id)
        if file.user_id != current_user.id:
            flash('Access denied')
            return redirect(url_for('core.dashboard'))
        
        # Get file data
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], file.filename)
        df = pd.read_csv(file_path) if file.file_type == 'csv' else pd.read_excel(file_path)
        
        # Get latest analysis
        analysis = Analysis.query.filter_by(file_id=file_id).order_by(Analysis.created_date.desc()).first()
        
        if not analysis:
            flash('No analysis found for this file')
            return redirect(url_for('core.view_file', file_id=file_id))
        
        # Use simpler Excel export if weasyprint is not available
        if not weasyprint_available:
            # Basic Excel export without ExportService
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='Data', index=False)
                
                # Add financial ratios if available
                if 'financial_ratios' in analysis.results:
                    ratios = pd.DataFrame(list(analysis.results['financial_ratios'].items()), 
                                         columns=['Ratio', 'Value'])
                    ratios.to_excel(writer, sheet_name='Ratios', index=False)
                
                # Apply simple formatting
                workbook = writer.book
                for worksheet in writer.sheets.values():
                    worksheet.set_column('A:Z', 15)
                    
            output.seek(0)
            
            return send_file(
                output,
                download_name=f'analysis_{file.filename}_{datetime.now():%Y%m%d}.xlsx',
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Full featured export with ExportService
            export_data = {
                'data': df.to_dict('records'),
                'ratios': analysis.results.get('financial_ratios', {}),
                'trends': analysis.results.get('trends', {}),
            }
            
            excel = ExportService.generate_excel_report(export_data, file.filename)
            
            return send_file(
                excel,
                download_name=f'analysis_{file.filename}_{datetime.now():%Y%m%d}.xlsx',
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}')
        return redirect(url_for('core.view_file', file_id=file_id))

@bp.route('/excel/comparison/<path:file_ids>')
@login_required
def export_comparison_excel(file_ids):
    try:
        file_id_list = [int(id) for id in file_ids.split(',')]
        files = []
        
        for file_id in file_id_list:
            file = FinancialFile.query.get_or_404(file_id)
            if file.user_id != current_user.id:
                flash('Access denied')
                return redirect(url_for('core.dashboard'))
            files.append(file)
            
        # Use simpler Excel export if weasyprint is not available
        if not weasyprint_available:
            # Basic comparison data export
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                for i, file in enumerate(files):
                    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], file.filename)
                    df = pd.read_csv(file_path) if file.file_type == 'csv' else pd.read_excel(file_path)
                    df.to_excel(writer, sheet_name=f'File {i+1}', index=False)
                    
                # Summary sheet
                summary = pd.DataFrame({
                    'Filename': [f.filename for f in files],
                    'Row Count': [pd.read_csv(os.path.join(current_app.config['UPLOAD_FOLDER'], f.filename)).shape[0] 
                                if f.file_type == 'csv' else 
                                pd.read_excel(os.path.join(current_app.config['UPLOAD_FOLDER'], f.filename)).shape[0] 
                                for f in files]
                })
                summary.to_excel(writer, sheet_name='Summary', index=False)
                
            output.seek(0)
            
            return send_file(
                output,
                download_name=f'comparison_{datetime.now():%Y%m%d}.xlsx',
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            # Full featured comparison with ExportService
            # Load data for comparison
            dataframes = []
            for file in files:
                file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], file.filename)
                df = pd.read_csv(file_path) if file.file_type == 'csv' else pd.read_excel(file_path)
                dataframes.append(df)
                
            # Get comparison data
            from app.comparison.service import ComparisonService
            comparison = ComparisonService(dataframes)
            comparison_data = {
                'summary_stats': comparison.compare_summary_statistics(),
                'differences': comparison.calculate_differences()
            }
            
            excel = ExportService.generate_comparison_excel(comparison_data, files)
            
            return send_file(
                excel,
                download_name=f'comparison_{datetime.now():%Y%m%d}.xlsx',
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
    except Exception as e:
        flash(f'Error generating comparison Excel: {str(e)}')
        return redirect(url_for('core.dashboard'))

@bp.route('/export/dashboard/<format>')
@login_required
def export_dashboard(format):
    """Export dashboard data in various formats"""
    # Get user's latest file for reference
    latest_file = FinancialFile.query.filter_by(user_id=current_user.id).order_by(
        FinancialFile.upload_date.desc()).first()
    
    filename_prefix = f"FinGenius_Dashboard_{datetime.now().strftime('%Y%m%d')}"
    
    if format == 'pdf':
        # Render HTML for PDF
        html = render_template('export/dashboard_report.html', 
                               user=current_user, 
                               timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"),
                               latest_file=latest_file)
        
        # Generate PDF using a helper function (needs to be defined or imported)
        try:
            from .service import generate_pdf
            pdf_file = generate_pdf(html)
            
            # Create response
            response = make_response(pdf_file)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename_prefix}.pdf"'
            return response
        except ImportError:
            return jsonify({"error": "PDF generation service not available"}), 500
    
    elif format == 'excel':
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Summary"
        
        # Add title
        ws['A1'] = "FinGenius Dashboard Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Add timestamp
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:F2')
        
        # Add headers
        headers = ["Metric", "Current Value", "Previous Value", "Change", "% Change"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add sample data (in a real app this would be actual data)
        metrics = [
            ["Revenue", 485726, 450000, 35726, "7.9%"],
            ["Profit", 142590, 135000, 7590, "5.6%"],
            ["ROI", "18.2%", "17.5%", "0.7%", "4.0%"],
            ["YoY Growth", "15.4%", "14.2%", "1.2%", "8.5%"]
        ]
        
        for row_idx, metric in enumerate(metrics, start=5):
            for col_idx, value in enumerate(metric, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to in-memory file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{filename_prefix}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    elif format == 'csv':
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(["FinGenius Dashboard Report"])
        writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        writer.writerow([])  # Empty row
        writer.writerow(["Metric", "Current Value", "Previous Value", "Change", "% Change"])
        
        # Write sample data (in a real app this would be actual data)
        metrics = [
            ["Revenue", 485726, 450000, 35726, "7.9%"],
            ["Profit", 142590, 135000, 7590, "5.6%"],
            ["ROI", "18.2%", "17.5%", "0.7%", "4.0%"],
            ["YoY Growth", "15.4%", "14.2%", "1.2%", "8.5%"]
        ]
        
        for metric in metrics:
            writer.writerow(metric)
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename_prefix}.csv"'
        return response
    
    return jsonify({"error": "Invalid export format"}), 400

@bp.route('/export/analysis/<int:analysis_id>/<format>')
@login_required
def export_analysis(analysis_id, format):
    """Export analysis results in various formats"""
    # Get the analysis
    analysis = Analysis.query.get_or_404(analysis_id)
    
    # Check if user owns the analysis
    file = FinancialFile.query.get_or_404(analysis.file_id)
    if file.user_id != current_user.id:
        return jsonify({"error": "Access denied"}), 403
    
    filename_prefix = f"FinGenius_Analysis_{datetime.now().strftime('%Y%m%d')}"
    
    if format == 'pdf':
        # Render HTML for PDF
        html = render_template('export/analysis_report.html', 
                               user=current_user, 
                               file=file,
                               analysis=analysis,
                               timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"))
        
        # Generate PDF
        pdf_file = generate_pdf(html)
        
        # Create response
        response = make_response(pdf_file)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename_prefix}.pdf"'
        return response
    
    elif format == 'excel':
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Analysis"
        
        # Add title
        ws['A1'] = f"Financial Analysis Report: {file.filename}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Add timestamp
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:F2')
        
        # Add analysis data
        row = 4
        if analysis.results:
            # Add financial ratios section
            if 'financial_ratios' in analysis.results:
                ws.cell(row=row, column=1, value="Financial Ratios")
                ws.cell(row=row, column=1).font = Font(size=14, bold=True)
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                ws.cell(row=row, column=1, value="Ratio")
                ws.cell(row=row, column=2, value="Value")
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).font = Font(bold=True)
                row += 1
                
                for ratio_name, ratio_value in analysis.results['financial_ratios'].items():
                    display_name = ' '.join(ratio_name.split('_')).title()
                    ws.cell(row=row, column=1, value=display_name)
                    ws.cell(row=row, column=2, value=ratio_value)
                    row += 1
                
                row += 1  # Add space
            
            # Add trend analysis section
            if 'trend_analysis' in analysis.results:
                ws.cell(row=row, column=1, value="Trend Analysis")
                ws.cell(row=row, column=1).font = Font(size=14, bold=True)
                ws.merge_cells(f'A{row}:F{row}')
                row += 1
                
                # Add trends
                for trend in analysis.results['trend_analysis']:
                    ws.cell(row=row, column=1, value=trend)
                    row += 1
                
                row += 1  # Add space
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save to in-memory file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{filename_prefix}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    elif format == 'csv':
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([f"Financial Analysis Report: {file.filename}"])
        writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        writer.writerow([])  # Empty row
        
        # Add analysis data
        if analysis.results:
            # Add financial ratios section
            if 'financial_ratios' in analysis.results:
                writer.writerow(["Financial Ratios"])
                writer.writerow(["Ratio", "Value"])
                
                for ratio_name, ratio_value in analysis.results['financial_ratios'].items():
                    display_name = ' '.join(ratio_name.split('_')).title()
                    writer.writerow([display_name, ratio_value])
                
                writer.writerow([])  # Empty row
            
            # Add trend analysis section
            if 'trend_analysis' in analysis.results:
                writer.writerow(["Trend Analysis"])
                
                # Add trends
                for trend in analysis.results['trend_analysis']:
                    writer.writerow([trend])
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename_prefix}.csv"'
        return response
    
    return jsonify({"error": "Invalid export format"}), 400

@bp.route('/export/projection/<format>')
@login_required
def export_projection(format):
    """Export projection data in various formats"""
    # Get parameters from query string
    projection_data = request.args.get('data')
    if not projection_data:
        return jsonify({"error": "No projection data provided"}), 400
    
    try:
        # Parse projection data
        import json
        data = json.loads(projection_data)
    except:
        return jsonify({"error": "Invalid projection data format"}), 400
    
    filename_prefix = f"FinGenius_Projection_{datetime.now().strftime('%Y%m%d')}"
    
    if format == 'pdf':
        # Render HTML for PDF
        html = render_template('export/projection_report.html', 
                               user=current_user, 
                               projection_data=data,
                               timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"))
        
        # Generate PDF
        try:
            from .service import generate_pdf
            pdf_file = generate_pdf(html)
            
            # Create response
            response = make_response(pdf_file)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename_prefix}.pdf"'
            return response
        except ImportError:
            return jsonify({"error": "PDF generation service not available"}), 500
    
    elif format == 'excel':
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Projections"
        
        # Add title
        ws['A1'] = "Financial Projections Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Add timestamp
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True)
        ws.merge_cells('A2:E2')
        
        # Add headers
        headers = ["Period", "Revenue", "Costs", "Profit", "Profit Margin"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add projection data
        row = 5
        for period_data in data:
            ws.cell(row=row, column=1, value=period_data['period'])
            ws.cell(row=row, column=2, value=period_data['revenue'])
            ws.cell(row=row, column=3, value=period_data['cost'])
            ws.cell(row=row, column=4, value=period_data['profit'])
            ws.cell(row=row, column=5, value=period_data['profitMargin'])
            row += 1
        
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to in-memory file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"{filename_prefix}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    elif format == 'csv':
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(["Financial Projections Report"])
        writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        writer.writerow([])  # Empty row
        writer.writerow(["Period", "Revenue", "Costs", "Profit", "Profit Margin"])
        
        # Add projection data
        for period_data in data:
            writer.writerow([
                period_data['period'],
                period_data['revenue'],
                period_data['cost'],
                period_data['profit'],
                period_data['profitMargin']
            ])
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename_prefix}.csv"'
        return response
    
    return jsonify({"error": "Invalid export format"}), 400

@bp.route('/export/chart/<format>')
@login_required
def export_chart(format):
    """Export a specific chart in various formats"""
    chart_type = request.args.get('type', 'revenue_expenses')
    
    if format == 'png':
        # For PNG, we'll handle this client-side with JavaScript
        return jsonify({"error": "PNG export is handled client-side"}), 400
    
    # Other formats would follow similar patterns to the above exports
    return jsonify({"error": "Not implemented yet"}), 501
